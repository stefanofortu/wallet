import shutil
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment, Font
from data.Results import Results
from data.CategoryStructure import CategoryStructure
from data.ExpenseGroups import ExpenseGroups
import logging
import os
import platform

logger = logging.getLogger("Stefano")


class ExcelWriter:
    def __init__(self, filename_in, template_sheetname, sheetname_title):
        if platform.system() == "Windows":
            path_separator = "\\"
        elif platform.system() == "Linux":
            path_separator = "/"
        else:
            path_separator = "\\"

        self.filename_in = os.getcwd() + path_separator + filename_in
        self.filename_out = os.getcwd() + path_separator + "output_files" + path_separator + self.create_output_name(filename_in)

        self.create_output_file()
        self.wb = openpyxl.load_workbook(self.filename_out)

        # OPTION 1 to create a new sheet from template
        # ws_template = self.wb.get_sheet_by_name(template_sheetname)
        # self.ws = self.wb.copy_worksheet(ws_template)
        # self.ws.title = self.sheetname

        # OPTION 2 to modify an existing wb
        self.ws = self.wb.get_sheet_by_name(template_sheetname)
        self.ws.title = sheetname_title
        self.ws.sheet_properties.tabColor = "FFFF00"

        self.move_sheet_tab_to_end()
        self.delete_all_other_sheets()

        self.ws.sheet_properties.outlinePr.summaryBelow = False

        # self.writer = pd.ExcelWriter(self.filename_in, engine='xlsxwriter')

    def create_output_name(self,filename_in):
        filename_in_no_extension = filename_in.split(".")[0]
        extension = filename_in.split(".")[1]
        base_name = filename_in_no_extension.split("_v")[0]
        version = int(filename_in_no_extension.split("_v")[1])
        filename = str(datetime.now().strftime("%Y%m%d_%Hh%Mm%Ss"))
        filename_out = base_name + "_v" + str(version).zfill(2) + "_" + filename + "." + extension
        return filename_out

    def create_output_file(self):
        try:
            os.makedirs(os.path.dirname(self.filename_out), exist_ok=True)
            shutil.copy(self.filename_in, self.filename_out)
        except shutil.SameFileError:
            logger.error("Source and destination represents the same file.")
        except PermissionError:
            logger.error("Permission denied.")
        #except:
            #logger.error("Error occurred while copying file.")
            

    def move_sheet_tab_to_end(self):
        sheets = self.wb.sheetnames
        # Remove the sheet from its current position
        sheets.remove(self.ws.title)

        # Insert the sheet at the last position
        sheets.append(self.ws.title)

        # Reorder the sheets in the workbook
        self.wb._sheets = [self.wb[sheet] for sheet in sheets]

        self.wb.active = self.wb.sheetnames.index(self.ws.title)

    def delete_all_other_sheets(self):
        for sheet_name in self.wb.sheetnames:
            if sheet_name != self.ws.title:
                sheet_to_delete = self.wb[sheet_name]
                self.wb.remove(sheet_to_delete)

    def __del__(self):
        self.wb.save(self.filename_out)
        self.wb.close()

    def write_main_category_results(self, main_category_results):
        if not isinstance(main_category_results, Results):
            raise TypeError("ExcelWriter.process(): Wrong input type for data main_category_results")

        category_column = 1
        in_column = 2
        savings_in_column = 3
        out_column = 4
        savings_out_column = 5
        no_tags_column = 6

        row_num = 130
        self.ws.cell(row_num, category_column).value = "Categories"
        self.ws.cell(row_num, in_column, "in")
        self.ws.cell(row_num, savings_in_column, "savings in")
        self.ws.cell(row_num, out_column, "out")
        self.ws.cell(row_num, savings_out_column, "savings out")
        self.ws.cell(row_num, no_tags_column, "no tags")

        for col in range(category_column, no_tags_column+1):
            self.ws.cell(row_num, col).alignment = Alignment(horizontal="center", vertical="center")
            self.ws.cell(row_num, col).font = Font(name='Calibri', size=11, color='FF000000', bold=True)

        row_num += 1
        block_row_num_start = row_num
        excel_structure = {"level1": [], "level2": []}

        for main_cat in list(CategoryStructure.categories.keys()):
            self.ws.cell(row_num, category_column, main_cat)
            self.ws.cell(row_num, in_column, main_category_results.df.loc[main_cat]["in"])
            self.ws.cell(row_num, savings_in_column, main_category_results.df.loc[main_cat]["savings_in"])
            self.ws.cell(row_num, out_column, main_category_results.df.loc[main_cat]["out"])
            self.ws.cell(row_num, savings_out_column, main_category_results.df.loc[main_cat]["savings_out"])
            self.ws.cell(row_num, no_tags_column, main_category_results.df.loc[main_cat]["no_tags"])
            row_num += 1
            start_group_level_1 = row_num
            for cat in list(CategoryStructure.categories[main_cat]):
                self.ws.cell(row_num, category_column, cat)
                self.ws.cell(row_num, in_column, main_category_results.df.loc[cat]["in"])
                self.ws.cell(row_num, savings_in_column, main_category_results.df.loc[cat]["savings_in"])
                self.ws.cell(row_num, out_column, main_category_results.df.loc[cat]["out"])
                self.ws.cell(row_num, savings_out_column, main_category_results.df.loc[cat]["savings_out"])
                self.ws.cell(row_num, no_tags_column, main_category_results.df.loc[cat]["no_tags"])
                end_group_level_1 = row_num
                row_num += 1
            excel_structure["level1"].append((start_group_level_1, end_group_level_1))

        block_row_num_end = row_num
        for t in excel_structure["level1"]:
            self.ws.row_dimensions.group(t[0], t[1], hidden=True, outline_level=1)

        for r_num in range(block_row_num_start, block_row_num_end):
            for col in range(category_column, no_tags_column+1):
                self.ws.cell(r_num, col).font = Font(name='Calibri', size=11, color='FF000000')

            for col in range(in_column, no_tags_column+1):
                self.ws.cell(r_num, col).number_format = "#,##0 [$€-2]"
                self.ws.cell(r_num, col).alignment = Alignment(horizontal="right", vertical="center")

        self.ws.cell(row_num, category_column, "-")
        self.ws.cell(row_num, in_column, "-")
        self.ws.cell(row_num, savings_in_column, "-")
        self.ws.cell(row_num, out_column, "-")
        self.ws.cell(row_num, savings_out_column, "-")

    def write_group_results(self, group_results):
        if not isinstance(group_results, Results):
            raise TypeError("ExcelWriter.process(): Wrong input type for data group_results")

        category_column = 1
        in_column = 2
        savings_in_column = 3
        out_column = 4
        savings_out_column = 5
        no_tags_column = 6
        excel_structure = {"level1": [], "level2": []}
        row_num = 50

        self.ws.cell(row_num, category_column).value = "Categories"
        self.ws.cell(row_num, in_column, "in")
        self.ws.cell(row_num, savings_in_column, "savings in")
        self.ws.cell(row_num, out_column, "out")
        self.ws.cell(row_num, savings_out_column, "savings out")
        self.ws.cell(row_num, no_tags_column, "no tags")

        for col in range(category_column, no_tags_column+1):
            self.ws.cell(row_num, col).alignment = Alignment(horizontal="center", vertical="center")
            self.ws.cell(row_num, col).font = Font(name='Calibri', size=11, color='FF000000', bold=True)
        row_num += 1
        block_row_num_start = row_num

        for main_group in list(ExpenseGroups.expense_groups.keys()):
            self.ws.cell(row_num, category_column, main_group)
            self.ws.cell(row_num, in_column, group_results.df.loc[main_group]["in"])
            self.ws.cell(row_num, savings_in_column, group_results.df.loc[main_group]["savings_in"])
            self.ws.cell(row_num, out_column, group_results.df.loc[main_group]["out"])
            self.ws.cell(row_num, savings_out_column, group_results.df.loc[main_group]["savings_out"])
            self.ws.cell(row_num, no_tags_column, group_results.df.loc[main_group]["no_tags"])
            row_num += 1
            start_group_level_1 = row_num
            for sub_group in ExpenseGroups.expense_groups[main_group].keys():
                self.ws.cell(row_num, category_column, sub_group)
                self.ws.cell(row_num, in_column, group_results.df.loc[sub_group]["in"])
                self.ws.cell(row_num, savings_in_column, group_results.df.loc[sub_group]["savings_in"])
                self.ws.cell(row_num, out_column, group_results.df.loc[sub_group]["out"])
                self.ws.cell(row_num, savings_out_column, group_results.df.loc[sub_group]["savings_out"])
                self.ws.cell(row_num, no_tags_column, group_results.df.loc[sub_group]["no_tags"])
                row_num += 1
                start_group_level_2 = row_num
                for cat in list(ExpenseGroups.expense_groups[main_group][sub_group]):
                    self.ws.cell(row_num, category_column, cat)
                    self.ws.cell(row_num, in_column, group_results.df.loc[cat]["in"])
                    self.ws.cell(row_num, savings_in_column, group_results.df.loc[cat]["savings_in"])
                    self.ws.cell(row_num, out_column, group_results.df.loc[cat]["out"])
                    self.ws.cell(row_num, savings_out_column, group_results.df.loc[cat]["savings_out"])
                    self.ws.cell(row_num, no_tags_column, group_results.df.loc[cat]["no_tags"])

                    end_group_level_1 = row_num
                    end_group_level_2 = row_num
                    row_num += 1
                excel_structure["level2"].append((start_group_level_2, end_group_level_2))
            excel_structure["level1"].append((start_group_level_1, end_group_level_1))

        block_row_num_end = row_num

        for t in excel_structure["level1"]:
            self.ws.row_dimensions.group(t[0], t[1], hidden=True, outline_level=1)
        for t in excel_structure["level2"]:
            self.ws.row_dimensions.group(t[0], t[1], hidden=True, outline_level=2)

        for r_num in range(block_row_num_start, block_row_num_end):
            for col in range(category_column, no_tags_column+1):
                self.ws.cell(r_num, col).font = Font(name='Calibri', size=11, color='FF000000')
                self.ws.cell(r_num, col).alignment = Alignment(horizontal="left", vertical="center")

            for col in range(in_column, no_tags_column+1):
                self.ws.cell(r_num, col).number_format = "#,##0 [$€-2]"
                self.ws.cell(r_num, col).alignment = Alignment(horizontal="right", vertical="center")

        if row_num > 130:
            logger.warning("ExcelWriter::write_group_results() - out of boundary box for group results results")
